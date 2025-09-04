import re
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
import os

rootdir = r"C:\微云同步助手\281475128776228\xinxipingtai\xinxipingtai.github.io"
inputfilename = "信息工程系人名单网页版.xlsx"
sheetname = "Sheet1"


inputfile = os.path.join(rootdir, inputfilename)
outputfilename = inputfilename + ".html"
outputfile = os.path.join(rootdir, outputfilename)
wb = load_workbook(filename=inputfile)
sheet = wb[sheetname]

cell_dic = {}
col_width = {}
row_height = {}

# 查询列宽
for col in sheet.columns:
    pat = r"[A-Z]+"
    pat = re.compile(pat)
    colname = pat.findall(col[0].coordinate)[0]
    px = round(sheet.column_dimensions[colname].width * 5)
    col_width[colname] = px

# 查询行高
for row in sheet.rows:
    pat = r"[A-Z]+(\d+)"
    pat = re.compile(pat)
    rowid = int(pat.findall(row[0].coordinate)[0])
    px = sheet.row_dimensions[rowid].height
    if px is None:
        px = 13.5
    row_height[str(rowid)] = px

# 找出所有合并区域的行高，列宽，向右合并距离，向下合并距离
for merged_range in sheet.merged_cells.ranges:
    now_width, now_height = 0, 0
    for i in range(merged_range.min_col, merged_range.max_col + 1): 
        coord = sheet.cell(row=1, column=i).coordinate
        colname = re.compile(r"[A-Z]+").findall(coord)[0]
        now_width += col_width[colname]
    for i in range(merged_range.min_row, merged_range.max_row + 1):
        coord = sheet.cell(row=i, column=1).coordinate
        colindex = re.compile(r"[A-Z]+(\d+)").findall(coord)[0]
        now_height += row_height[colindex]
    cell_dic[sheet.cell(row=merged_range.min_row, column=merged_range.min_col)] = (now_height, now_width, merged_range.max_col - merged_range.min_col + 1, merged_range.max_row - merged_range.min_row + 1)

html = '<table style="border:1px solid #000000; border-collapse: collapse;width:100%;" border="1">'
for row in sheet.rows:
    tr = '<tr>'
    for cell in row:
        td = ""
        text = '' if cell.value is None else str(cell.value)
        text = text.replace("\n", "<br/>")
        print(text)
        if cell in cell_dic:
            vertical = f'vertical-align: {cell.alignment.vertical};' if cell.alignment.vertical else ''
            horizontal = f'text-align: {cell.alignment.horizontal};' if cell.alignment.horizontal else ''
            font_size = str(int(cell.font.size) + 3)
            font_weight = '700' if cell.font.b else '400'
            style = f'"color: rgb(0, 0, 0); font-size: {font_size}px; font-weight: {font_weight}; font-style: normal;{vertical}{horizontal}"'
            td = f'<td height="{cell_dic[cell][0]}" width="{cell_dic[cell][1]}" colspan="{cell_dic[cell][2]}" rowspan="{cell_dic[cell][3]}" style={style}>{text}</td>'
        elif not isinstance(cell, MergedCell):
            vertical = f'vertical-align: {cell.alignment.vertical};' if cell.alignment.vertical else ''
            horizontal = f'text-align: {cell.alignment.horizontal};' if cell.alignment.horizontal else ''
            cell_name, cell_index = re.compile(r"([A-Z]+)(\d+)").findall(cell.coordinate)[0]
            font_size = str(int(cell.font.size) + 3)
            font_weight = '700' if cell.font.b else '400'
            style = f'"color: rgb(0, 0, 0); font-size: {font_size}px; font-weight: {font_weight}; font-style: normal;{vertical}{horizontal}"'
            td = f'<td height="{row_height[cell_index]}" width="{col_width[cell_name]}" style={style}>{text}</td>' if text else f'<td height="{row_height[cell_index]}" width="{col_width[cell_name]}"></td>'
        tr += td
    tr += '</tr>'
    html += tr
html += '</table>'

with open(outputfile, 'w', encoding='utf-8') as f:
    f.write(html)
